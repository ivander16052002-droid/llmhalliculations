from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from src.config import PROJECT_ROOT


ANNOTATION_COLUMNS = [
    "svr",
    "fr",
    "mr",
    "mr_author",
    "mr_title",
    "mr_year",
    "mr_container",
    "qvr",
    "esr",
    "lvr",
    "source_language",
    "verification_link",
    "comment",
]


BASE_COLUMNS = [
    "run_id",
    "question_id",
    "discipline",
    "question",
    "source_ref",
    "model",
    "prompt_type",
    "status",
    "answer",
    "source_id",
    "source_authors",
    "source_title",
    "source_year",
    "source_container",
    "source_locator",
    "locator_present",
    "quote_id",
    "quote_text",
    "quote_present",
    "limitations",
    "parse_error",
]


def add_helper_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Add technical helper columns for easier annotation."""
    df = df.copy()

    for column in BASE_COLUMNS:
        if column not in df.columns:
            df[column] = ""

    df["source_present"] = (
        df["source_authors"].fillna("").astype(str).str.strip().ne("")
        | df["source_title"].fillna("").astype(str).str.strip().ne("")
        | df["source_year"].fillna("").astype(str).str.strip().ne("")
    )

    df["source_reference"] = (
        df["source_authors"].fillna("").astype(str).str.strip()
        + ". "
        + df["source_title"].fillna("").astype(str).str.strip()
        + ". "
        + df["source_year"].fillna("").astype(str).str.strip()
    ).str.replace(r"\s+", " ", regex=True).str.strip()

    df["quote_present"] = (
        df["quote_text"]
        .fillna("")
        .astype(str)
        .str.strip()
        .ne("")
    )

    df.loc[~df["source_present"], "source_reference"] = ""

    return df


def build_annotation_df(parsed_df: pd.DataFrame) -> pd.DataFrame:
    df = add_helper_columns(parsed_df)

    output_columns = [
        "run_id",
        "question_id",
        "discipline",
        "question",
        "source_ref",
        "model",
        "prompt_type",
        "status",
        "answer",
        "source_present",
        "source_reference",
        "source_id",
        "source_authors",
        "source_title",
        "source_year",
        "source_container",
        "source_locator",
        "locator_present",
        "quote_id",
        "quote_text",
        "quote_present",
        "limitations",
        "parse_error",
    ]

    annotation_df = df[output_columns].copy()

    for column in ANNOTATION_COLUMNS:
        annotation_df[column] = ""

    return annotation_df


def save_xlsx(df: pd.DataFrame, output_path: Path) -> None:
    df.to_excel(output_path, index=False)

    workbook = load_workbook(output_path)
    worksheet = workbook.active
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    widths = {
        "A": 24,  # run_id
        "B": 12,  # question_id
        "C": 28,  # discipline
        "D": 55,  # question
        "E": 12,  # source_ref
        "F": 14,  # model
        "G": 16,  # prompt_type
        "H": 12,  # status
        "I": 65,  # answer
        "J": 65,  # claim_text
        "K": 14,  # source_present
        "L": 50,  # source_reference
        "M": 12,  # source_id
        "N": 24,  # authors
        "O": 42,  # title
        "P": 12,  # year
        "Q": 35,  # container
        "R": 25,  # locator
        "S": 15,  # locator_present
        "T": 45,  # limitations
        "U": 25,  # parse_error
    }

    for column_letter, width in widths.items():
        worksheet.column_dimensions[column_letter].width = width

    # Annotation columns
    for col_idx in range(22, len(df.columns) + 1):
        worksheet.column_dimensions[get_column_letter(col_idx)].width = 16

    workbook.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--input-file",
        type=str,
        required=True,
        help="Parsed CSV file from data/parsed_responses.",
    )
    parser.add_argument(
        "--output-file",
        type=str,
        required=True,
        help="Output annotation file, .csv or .xlsx.",
    )
    args = parser.parse_args()

    input_path = PROJECT_ROOT / args.input_file
    output_path = PROJECT_ROOT / args.output_file
    output_path.parent.mkdir(parents=True, exist_ok=True)

    parsed_df = pd.read_csv(input_path, encoding="utf-8-sig")
    annotation_df = build_annotation_df(parsed_df)

    if output_path.suffix.lower() == ".xlsx":
        save_xlsx(annotation_df, output_path)
    else:
        annotation_df.to_csv(output_path, index=False, encoding="utf-8-sig")

    print(f"Annotation rows: {len(annotation_df)}")
    print(f"Saved to: {output_path}")


if __name__ == "__main__":
    main()